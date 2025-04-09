from openpyxl import Workbook, load_workbook
wb = load_workbook("algas.xlsx")
ws = wb.active
max_col = ws.max_column
max_row = ws.max_row
print("Likme\tStundas\tKopā")
for r in range(2, max_row+1):
    likme = float(ws['B'+str(r)].value)
    stundas = int(ws['C'+str(r)].value)
    kopa = likme * stundas
    print(f"{likme} \t{stundas} \t{kopa}")